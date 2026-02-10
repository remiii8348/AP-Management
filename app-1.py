import streamlit as st
from streamlit_gsheets import GSheetsConnection
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- [1] 페이지 설정 및 보안 (와이드 레이아웃) ---
st.set_page_config(page_title="미지급금 통합 관리 시스템", layout="wide")

st.markdown("""
    <style>
        .block-container { padding-top: 1.5rem; max-width: 98%; }
        .stTabs [data-baseweb="tab-list"] { gap: 24px; }
        .stTabs [data-baseweb="tab"] { height: 50px; font-size: 18px; font-weight: bold; }
    </style>
""", unsafe_allow_html=True)

def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False
    if "password_correct" not in st.session_state:
        _, col, _ = st.columns([1, 2, 1])
        with col: st.text_input("🔑 관리자 비밀번호", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        _, col, _ = st.columns([1, 2, 1])
        with col:
            st.text_input("🔑 관리자 비밀번호", type="password", on_change=password_entered, key="password")
            st.error("😕 비밀번호가 틀렸습니다.")
        return False
    return True

# --- [2] 메인 앱 로직 시작 ---
if check_password():
    conn = st.connection("gsheets", type=GSheetsConnection)

    # 데이터 로딩 함수 (날짜 에러 및 빈 데이터 방어 로직 포함)
    def load_full_data():
        # 메인 데이터 시트 로드
        main_df = conn.read(worksheet="시트1", ttl=0)
        main_df['Date'] = pd.to_datetime(main_df['Date'], errors='coerce')
        main_df = main_df.dropna(subset=['Date'])
        main_df['Amount_KRW'] = pd.to_numeric(main_df['Amount_KRW'], errors='coerce').fillna(0).astype(int)
        
        # 메모장 시트 로드
        try:
            notes_df = conn.read(worksheet="special_notes", ttl=0)
        except:
            notes_df = pd.DataFrame(columns=['Content'])
        return main_df, notes_df

    # 엑셀 내보내기 (테두리, 가운데정렬, 10pt, SUM 함수 적용)
    def convert_to_excel(df_export):
        output = io.BytesIO()
        exp = df_export[['Date', 'Vendor', 'Amount_KRW']].copy()
        exp['Date'] = exp['Date'].dt.strftime('%Y-%m-%d')
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            exp.to_excel(writer, index=False, sheet_name='미지급목록')
            ws = writer.sheets['미지급목록']
            
            # 스타일 설정
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="D9EAD3", fill_type="solid")
            sum_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
            font_style = Font(name='맑은 고딕', size=10)
            
            # 모든 셀 적용
            for row in ws.iter_rows(min_row=1, max_row=len(exp)+1, min_col=1, max_col=3):
                for cell in row:
                    cell.font = font_style
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.row == 1: cell.fill = header_fill
                    if cell.column == 3 and cell.row > 1: cell.number_format = '#,##0'

            # 합계 행 및 SUM 함수 적용
            last_r = len(exp) + 2
            ws.cell(row=last_r, column=1, value="합계").fill = sum_fill
            ws.cell(row=last_r, column=1).font = Font(name='맑은 고딕', size=10, bold=True)
            ws.cell(row=last_r, column=2, value="").fill = sum_fill
            
            sum_cell = ws.cell(row=last_r, column=3, value=f"=SUM(C2:C{last_r-1})")
            sum_cell.fill = sum_fill
            sum_cell.font = Font(name='맑은 고딕', size=10, bold=True, color="0000FF")
            sum_cell.number_format = '#,##0'

            # 열 너비 자동 조절
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
        return output.getvalue()

    # 데이터 가져오기
    df, notes_df = load_full_data()
    st.title("💸 미지급금 통합 관리 시스템")
    
    # 탭 구성 (일괄 업로드 삭제)
    tab1, tab2 = st.tabs(["📋 미지급 관리 & 메모", "🔍 히스토리 조회 & 수정"])

    # --- Tab 1: 미지급 관리 ---
    with tab1:
        with st.form("in_form", clear_on_submit=True):
            st.subheader("📝 신규 내역 입력")
            f1, f2, f3, f4, f5, f6 = st.columns([1, 2, 0.8, 1.2, 1, 1])
            with f1: in_date = st.date_input("지급날짜", datetime.now())
            with f2: in_vendor = st.text_input("거래처명")
            with f3: in_curr = st.selectbox("통화", ["KRW", "USD", "AUD"])
            with f4: in_amt = st.number_input("금액", min_value=0.0)
            with f5: in_rate = st.number_input("환율", min_value=1.0, value=1350.0 if in_curr == "USD" else 1.0)
            with f6: st.write(""); in_fixed = st.checkbox("고정지출(1년)")
            if st.form_submit_button("➕ 추가", use_container_width=True):
                if in_vendor:
                    count = 12 if in_fixed else 1
                    new_rows = []
                    for i in range(count):
                        d = pd.to_datetime(in_date) + pd.DateOffset(months=i)
                        new_rows.append({'Date': d, 'Vendor': in_vendor, 'Currency': in_curr, 'Amount_F': in_amt, 'Ex_Rate': in_rate, 'Amount_KRW': int(in_amt*in_rate), 'Status': 'Wait', 'Is_Fixed': in_fixed})
                    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
                    conn.update(worksheet="시트1", data=df); st.rerun()

        st.divider()
        st.subheader("📌 특이사항 메모 (체크리스트)")
        n1, n2 = st.columns([6, 1])
        with n1: note_txt = st.text_input("업무 특이사항 입력", placeholder="예: 체리 파손 건 보험 청구 확인 필요")
        with n2: st.write(""); 
            if st.button("메모 추가", use_container_width=True):
                if note_txt:
                    notes_df = pd.concat([notes_df, pd.DataFrame([{'Content': note_txt}])], ignore_index=True)
                    conn.update(worksheet="special_notes", data=notes_df); st.rerun()
        
        if not notes_df.empty:
            for idx, row in notes_df.iterrows():
                nc1, nc2 = st.columns([6, 1])
                nc1.info(row['Content'])
                if nc2.button("완료", key=f"nt_{idx}"):
                    notes_df = notes_df.drop(idx); conn.update(worksheet="special_notes", data=notes_df); st.rerun()

        st.divider()
        st.subheader("🔍 기간별 미지급 조회")
        unpaid_only = df[df['Status'] == 'Wait']
        oldest_d = pd.to_datetime(unpaid_only['Date']).min().date() if not unpaid_only.empty else datetime.now().date()
        c1, c2, c3 = st.columns([1.5, 1.5, 2])
        with c1: start_d = st.date_input("조회 시작", oldest_d)
        with c2: end_d = st.date_input("조회 종료", datetime.now().date() + timedelta(days=14))
        
        view_df = df[(df['Date'].dt.date >= start_d) & (df['Date'].dt.date <= end_d) & (df['Status'] == 'Wait')].sort_values('Date')
        
        with c3: st.write(""); 
            if not view_df.empty:
                st.download_button("📥 현재 조회내역 엑셀 다운로드", data=convert_to_excel(view_df), file_name=f"AP_Report_{datetime.now().strftime('%m%d')}.xlsx", use_container_width=True)

        if not view_df.empty:
            v0, v1, v2, v3, v4 = st.columns([0.5, 1.2, 2.5, 4, 1])
            v0.write("**삭제**"); v1.write("**날짜**"); v2.write("**거래처**"); v3.write("**금액**"); v4.write("**완료**")
            today = datetime.now().date()
            for idx, row in view_df.iterrows():
                r0, r1, r2, r3, r4 = st.columns([0.5, 1.2, 2.5, 4, 1])
                if r0.button("🗑️", key=f"d_{idx}"):
                    df = df.drop(idx); conn.update(worksheet="시트1", data=df); st.rerun()
                d_val = row['Date'].date()
                d_str = d_val.strftime('%Y-%m-%d')
                if d_val == today: r1.write(f":green-background[**{d_str}**]")
                elif d_val < today: r1.write(f":red[**{d_str}**]")
                else: r1.write(f"**{d_str}**")
                r2.write(f"**{row['Vendor']}**")
                r3.write(f"**{int(row['Amount_KRW']):,} 원**" + (f" ({row['Amount_F']:,.1f}{row['Currency']})" if row['Currency']!='KRW' else ""))
                if r4.button("✅", key=f"p_{idx}"):
                    df.at[idx, 'Status'] = 'Done'; conn.update(worksheet="시트1", data=df); st.rerun()
            
            st.divider()
            _, s2, s3 = st.columns([3, 1, 3])
            s2.write("### 합계")
            s3.write(f"### :blue[{int(view_df['Amount_KRW'].sum()):,} 원]")

    # --- Tab 2: 히스토리 조회 및 수정 ---
    with tab2:
        st.subheader("🔎 히스토리 필터 및 상세 수정")
        
        s_col1, s_col2 = st.columns(2)
        with s_col1: search_cat = st.radio("상태 필터", ["미지급(Wait)", "지급완료(Done)", "전체"], horizontal=True)
        with s_col2: 
            v_list = ["전체"] + sorted(df['Vendor'].unique().tolist())
            search_v = st.selectbox("거래처별 조회", v_list)
        
        h_df = df.copy()
        if search_cat == "미지급(Wait)": h_df = h_df[h_df['Status'] == 'Wait']
        elif search_cat == "지급완료(Done)": h_df = h_df[h_df['Status'] == 'Done']
        if search_v != "전체": h_df = h_df[h_df['Vendor'] == search_v]
        
        st.write(f"📊 검색 결과: {len(h_df)}건")
        
        # 엑셀 내보내기 버튼 (히스토리 탭 전용)
        if not h_df.empty:
            st.download_button(f"📥 {search_v} 필터 결과 엑셀 내보내기", data=convert_to_excel(h_df), file_name=f"History_{search_v}.xlsx")
            
            # 데이터 에디터 (수정 가능)
            edited = st.data_editor(h_df.sort_values('Date', ascending=False), use_container_width=True, hide_index=True)
            if st.button("💾 위 수정사항 구글 시트에 최종 저장"):
                edited['Amount_KRW'] = (edited['Amount_F'] * edited['Ex_Rate']).astype(int)
                # 전체 데이터프레임 업데이트 로직
                df.set_index(df.index, inplace=True)
                df.update(edited)
                conn.update(worksheet="시트1", data=df)
                st.success("데이터가 안전하게 저장되었습니다!"); st.rerun()